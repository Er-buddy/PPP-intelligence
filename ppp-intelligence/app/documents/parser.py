from pathlib import Path

import pymupdf
from docx import Document as DocxDocument
from openpyxl import load_workbook


def parse_pdf(path: Path) -> list[dict]:
    pages = []

    with pymupdf.open(path) as document:
        for index, page in enumerate(document):
            text = page.get_text("text").strip()

            if not text:
                continue

            pages.append(
                {
                    "page": index + 1,
                    "text": text,
                    "metadata": {
                        "source_type": "pdf",
                    },
                }
            )

    if not pages:
        raise ValueError(
            "No machine-readable text found. "
            "OCR is not included in the MVP."
        )

    return pages


def parse_docx(path: Path) -> list[dict]:
    document = DocxDocument(path)

    paragraphs = [
        paragraph.text.strip()
        for paragraph in document.paragraphs
        if paragraph.text.strip()
    ]

    if not paragraphs:
        raise ValueError("No readable text found in DOCX.")

    return [
        {
            "page": None,
            "text": "\n".join(paragraphs),
            "metadata": {
                "source_type": "docx",
            },
        }
    ]


def parse_xlsx(path: Path) -> list[dict]:
    workbook = load_workbook(
        path,
        read_only=True,
        data_only=True,
    )

    pages = []

    for sheet in workbook.worksheets:
        rows = []

        for row in sheet.iter_rows(values_only=True):
            values = [
                str(value).strip()
                for value in row
                if value is not None
            ]

            if values:
                rows.append(" | ".join(values))

        if rows:
            pages.append(
                {
                    "page": None,
                    "text": "\n".join(rows),
                    "metadata": {
                        "source_type": "xlsx",
                        "sheet": sheet.title,
                    },
                }
            )

    if not pages:
        raise ValueError("No readable content found in XLSX.")

    return pages


def parse_document(path: Path) -> list[dict]:
    extension = path.suffix.lower()

    if extension == ".pdf":
        return parse_pdf(path)

    if extension == ".docx":
        return parse_docx(path)

    if extension == ".xlsx":
        return parse_xlsx(path)

    raise ValueError(f"Unsupported file type: {extension}")
